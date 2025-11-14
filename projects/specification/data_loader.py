from projects.specification.tables_config import TableConfigInventor

from openpyxl import load_workbook
from collections import OrderedDict


def __get_dict_from_xl(filepath: str, inventor_table: TableConfigInventor) -> list[list]:
    book = load_workbook(filepath)
    sheet = book.active

    dict_name_number_col: OrderedDict[str, int] = OrderedDict()
    for col_number in range(1, sheet.max_column + 1):
        col_value = sheet.cell(1, col_number).value
        if col_value in inventor_table.get_columns_name():
            dict_name_number_col[col_value] = col_number

    dict_data = OrderedDict()
    for row in range(2, sheet.max_row + 1):
        column_name = inventor_table.get_columns_name()
        number_key_column_name = [dict_name_number_col[column_name[i]] for i in inventor_table.index_keys]
        key = tuple(sheet.cell(row=row, column=col).value for col in number_key_column_name)
        key = tuple('' if k is None else k for k in key)

        column_name = inventor_table.get_columns_name()[inventor_table.index_values[0]]
        count_value = str(sheet.cell(row=row, column=dict_name_number_col[column_name]).value).split()

        if len(count_value) > 1:
            count, unit = count_value
        else:
            count, unit = count_value[0], 'шт.'
        count = float(count.replace(',', '.'))    
        value = {
            'count': count,
            'unit': unit
            }

        if key in dict_data:
            dict_data[key]['count'] += count
        else:
            dict_data[key] = value

    return dict_data

def __dict_data_to_2list(dict_data: OrderedDict[tuple[str], dict[float, str]], inventor_table: TableConfigInventor) -> list[list]:    
    list_data = []
    for key, value in dict_data.items():
        list_item = list(key) + list(value.values())
        list_index = inventor_table.index_keys + inventor_table.index_values

        tpl = tuple((index, item) for index, item in zip(list_index, list_item))
        list_item_sort = sorted(tpl, key=lambda x: x[0])

        list_row = []
        for _, item in sorted(list_item_sort):
            list_row.append(item)
        list_data.append(list_row)
    return list_data

def get_data_from_xl(filepath: str, inventor_table_config: TableConfigInventor) -> tuple[TableConfigInventor, list[list]]:
    dict_xl = __get_dict_from_xl(filepath=filepath, inventor_table=inventor_table_config)
    return __dict_data_to_2list(dict_data=dict_xl, inventor_table=inventor_table_config)


if __name__ == '__main__':
    p = r'D:\Python\AlfaServis\Constructor\projects\specification\DEBUG\ALS.1642.4.2.01.00.00.000 СБ - нивентор.xlsx'
    inventor_table_config = TableConfigInventor()
    data = get_data_from_xl(p, inventor_table_config)
    print(data)