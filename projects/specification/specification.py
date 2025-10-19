import openpyxl
from openpyxl.utils.cell import coordinate_from_string
import os
import pathlib


dct_cols_name = {
    "Артикул": "Инвентарный номер",
    "Наименование": "Описание",
    "Технические характеристики": ["Обозначение", "Технические характеристики"],
    "КОЛ.": "Общ. кол-во, ",
    "Материал": "Материал"
}


cols_spec_from_inv = ("Инвентарный номер", "Описание", "Обозначение", "Технические характеристики", "КОЛ.", "Материал")
cols_spec_from_excel = ("Артикул", "Наименование", "Технические характеристики", "Общ. кол-во, \nшт.", "Ед. изм.", "Материал")
standart_armature = [
    "ободок", "clamp", "уплотнительное", "эллиптическая", "Переход",  "Отвод", "штуцер", "Гайка", "Тройник", "Труба",
    "Соединение", "Ниппель", "кольцо", "Шпилька", "проволочный", "Уголок", "Держатель", "Муфта", "пластина", "Монтажная",
    "Профиль", "Вварыш", "опор", "резьб", "Болт", "Переходник", "быстроразъемный", "Круг", "ГОСТ", "DIN", "ISO", "Винт"
]


def get_number_cols(sheet, cols: tuple, range_: str) -> list:
    number_cols = []
    for col_name in cols:
        for col in sheet[range_]:
            for cell in col:
                value = cell.value
                if value == col_name:
                    number_cols.append(coordinate_from_string(cell.coordinate)[0])
    return number_cols


def add_cat(row: list) -> list:
    for cell in row:
        for w in standart_armature:
            if w.lower() in cell.lower():
                row.append('standart')
                return row
    else:
        row.append('unknow')
    return row


def get_specification(path: str) -> list:
    book: openpyxl.Workbook = openpyxl.load_workbook(path)
    sheet = book['Спецификация']

    cols = get_number_cols(sheet, cols_spec_from_inv, 'A1:L1')

    dict_spec = {}
    for row in zip(*[sheet[f'{c}2:{c}{sheet.max_row}'] for c in cols]):
        art, des, name, char, q, material = tuple(i[0].value for i in row)

        key = (art, des, char if char else name, material)
        if q is not None and 'м' in str(q).lower():
            q = float(str(q).split()[0].replace(',', '.'))

        if key not in dict_spec:
            dict_spec[key] = q
        else:
            dict_spec[key] += q
    book.close()

    list_spec = []
    for key, value in dict_spec.items():
        row = list(map(str, (*key, value)))
        row[-1], row[-2] = row[-2], row[-1]
        if '.' in row[-2] or ',' in row[-2]:
            row.insert(-1, 'мм')
        else:
            row.insert(-1, 'шт')

        row = add_cat(row)
        list_spec.append(row)

    return sorted(list_spec, key=lambda r: (r[-1], r[1], r[2]))


def set_specification(path: str):
    spec = get_specification(path)

    book: openpyxl.Workbook = openpyxl.load_workbook(os.path.join(PROJECT_PATH,'ШАБЛОН. Закупка К.О.xlsm'), keep_vba=True, read_only=False)
    sheet = book['Закупка']
    # book: openpyxl.Workbook = openpyxl.load_workbook(path)
    # sheet = book['Спецификация']
    cols = get_number_cols(sheet, cols_spec_from_excel, 'A6:L6')

    for y, row in enumerate(spec, 8):
        for x, cell in enumerate(row[:-1]):
            cell = cell if cell != "None" else ""
            sheet[f'{cols[x]}{y}'] = cell

    book.save(os.path.join(PROJECT_PATH,'31 ШАБЛОН. Закупка К.О.xlsm'))
    book.close()


if __name__ == '__main__':
    PROJECT_PATH = pathlib.Path(__file__).parent
    p = os.path.join(PROJECT_PATH, r'ALS.1642.4.2.01.00.00.000 СБ - нивентор.xlsx')
    set_specification(path=p)
