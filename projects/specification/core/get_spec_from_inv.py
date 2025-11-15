import win32com.client as wc32
import openpyxl
import os
import pathlib
from datetime import datetime


def get_spec_from_inv() -> tuple:
    app = wc32.GetActiveObject("Inventor.Application")
    doc = app.ActiveDocument

    name = pathlib.Path(doc.File.FullFileName).stem
    date = f'{datetime.today():%d.%m.%Y %H-%M}'
    filename = f'{name}__inventor__{date}'
    full_filename = os.path.join(str(pathlib.Path(__file__).parent),f'{filename}.xlsx')
    print(full_filename)
    doc.ComponentDefinition.BOM.BOMViews[1].Export(full_filename, 74498)

    return filename, full_filename, date


def get_spec_assembly(document, is_recursion=False, dct=None) -> dict:
    if dct is None:
        dct = {}

    if not is_recursion:
        # Начало рекурсии. Когда мы работаем с базовым файлом
        list_component = document.ComponentDefinition.Occurrences
    else:
        # Рекурсия, когда уже начали перебор подсборок
        list_component = document.SubOccurrences

    for component in list_component:
        try:
            name = str(component.Name)
            key = (name, )
            value = ''

            count = int(component.SubOccurrences.Count)
            if count != 0:
                # Если компонентов в подсборке более 0, то тогда мы заходим в неё и перебираем в рекурсии
                sub_dct = get_spec_assembly(document=component, is_recursion=True, dct=dct)
                # for key, value in sub_dct:
                #     if key in dct:
                #         dct[key] += value
                
                # dct = sub_dct
            else:
                for prop in component.Definition.Document.PropertySets:
                    for p in prop:
                        try:
                            print(p.Value)
                        except Exception:
                            pass
        except Exception:
            pass
    return dct


def merge_spec_from_inv(full_filename: str, date: str):
    full_filename_to = full_filename.replace(f'__{date}')

    for filename in os.listdir():
        if filename in full_filename:
            
            book_to: openpyxl.Workbook = openpyxl.load_workbook(full_filename_to)
            sheet_to = book_to.create_sheet(date)

            book_from: openpyxl.Workbook = openpyxl.load_workbook(full_filename)
            sheet_from = book_from.active
            
            for row in sheet_from.iter_rows():
                for cell in row:
                    sheet_to[cell.coordinate].value = cell.value
            
            book_to.save(full_filename_to)
            book_to.close()
            book_to.close()
    else:
        book: openpyxl.Workbook = openpyxl.load_workbook(full_filename)
        sheet = book.active
        sheet.title = date
        os.rename(full_filename, full_filename_to)
        
    


def add_sheet(path_source: str, path_target: str, sheet_name: str):
    book_source: openpyxl.Workbook = openpyxl.load_workbook(path_source)
    sheet_to = book_source.create_sheet(sheet_name)

    book_target: openpyxl.Workbook = openpyxl.load_workbook(path_target)
    sheet_from = book_target.active

    for row in sheet_from.iter_rows():
        for cell in row:
            sheet_to[cell.coordinate].value = cell.value

    book_source.save(path_source)
    book_source.close()
    book_target.close()


def main():
    # filename, full_filename, date = get_spec_from_inv()
    # print(get_spec_from_inv())
    app = wc32.GetActiveObject('Inventor.Application')
    get_spec_assembly(document=app.ActiveDocument)
    # spec_from_inv = r'ALS.MRU35000.00.00.000.AL СБ_25.09.2024 17-37.xlsx'
    # n = '25.09.2024 17-18'

    # add_sheet(path_source=full_filename, path_target=full_filename, sheet_name=date)



if __name__ == '__main__':
    main()