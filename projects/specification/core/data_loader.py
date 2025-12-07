from openpyxl import load_workbook
from collections import OrderedDict


if __name__ == '__main__':
    import sys
    # Для запуска через IDE
    from pathlib import Path
    test_path = str(Path(__file__).parent.parent.parent.parent)
    sys.path.append(test_path)


from projects.specification.core.config_table import (
    ColumnConfig,
    GENERAL_ITEM_CONFIG,
    INVENTOR_ITEM_CONFIG,
)


def __get_dict_from_xlsx(filepath: str) -> dict[tuple[str,...], list[float, str]]:
    config_columns: list[ColumnConfig] = GENERAL_ITEM_CONFIG.columns + INVENTOR_ITEM_CONFIG.columns
    name_columns: list[str] = tuple(col.column_name for col in config_columns if col.is_view)
    key_name_columns: list[str] = [col.column_name for col in config_columns if col.is_key]

    book = load_workbook(filepath)
    sheet = book.active

    dict_name_number_col: dict[str, int] = {}
    for col_number in range(1, sheet.max_column + 1):
        col_value = sheet.cell(1, col_number).value
        if col_value in name_columns:
            dict_name_number_col[col_value] = col_number

    dict_data: dict[tuple[str,...], list[float, str]] = {}
    number_key_column_name = [dict_name_number_col[key_name] for key_name in key_name_columns]
    number_quantity, *_ = [dict_name_number_col[col.column_name] for col in config_columns if col.field == 'quantity']

    for row_number in range(2, sheet.max_row + 1):
        key = []
        for col_number in number_key_column_name:
            cell_value = sheet.cell(row=row_number, column=col_number).value
            cell_value = '' if cell_value is None else cell_value
            key.append(cell_value)
        key = tuple(key)

        quantity_value = str(sheet.cell(row=row_number, column=number_quantity).value).split()

        if len(quantity_value) > 1:
            count, unit = quantity_value
        else:
            count, unit = quantity_value[0], 'шт.'
        count = float(count.replace(',', '.'))    
        value = [count, unit]

        if key in dict_data:
            dict_data[key][0] += count
        else:
            dict_data[key] = value
    
    book.close()
    
    return dict_data

def __dict_data_to_2list(dict_data: OrderedDict[tuple[str], list[float, str]]) -> list[list]:   
    config_tuple: list[ColumnConfig] = GENERAL_ITEM_CONFIG.columns + INVENTOR_ITEM_CONFIG.columns
        
    key_indexs = []
    value_indexs = []
    for i, col in enumerate(config_tuple):
        if col.is_view:
            if col.is_key:
                key_indexs.append(i)
            elif col.is_value:
                value_indexs.append(i)    

    list_data = []
    for number, (key, value) in enumerate(dict_data.items(), 1):
        row = [None] * len(config_tuple)

        for i, v in zip(key_indexs + value_indexs, key + tuple(value)):
            row[i] = v

        list_data.append(row)
    return list_data

def get_specifitaction_inventor_from_xlsx(filepath: str) -> list[list]:
    dict_xl = __get_dict_from_xlsx(filepath=filepath)
    return __dict_data_to_2list(dict_data=dict_xl)


if __name__ == '__main__':
    p = r'D:\Python\AlfaServis\Constructor\projects\specification\DEBUG\ALS.1642.5.3.05.01.00.000 - Инвентор.xlsx'
    data = get_specifitaction_inventor_from_xlsx(filepath=p)
    
    for row in data:
        print(row)

