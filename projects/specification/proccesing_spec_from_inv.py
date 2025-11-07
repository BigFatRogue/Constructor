
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime


COL_NAMES_PROPERTY = {
    'A': 'Инвентарный номер', 
    'B': 'Описание', 
    'E': 'Материал', 
    'F': 'Обозначение', 
    'G': 'Технические характеристики'
    }
COL_COUNT = 'C'
COL_GROUP = 'H'


def get_dict_from_sheet(sheet, dict_sheet_after:dict=None) -> dict:
    """
    Получение dict из лист, где одинаковые позиции суммируются\n
    dict[key] = value\n
    key = tuple(COL_NAMES_PROPERTY)\n
    value = list(COL_COUNT, COL_GROUP)\n
    returt: dict
    """
    dict_current_sheet = {}
    for y in range(2, 1000):
        key = tuple(sheet[f'{col_adrs}{y}'].value for col_adrs in COL_NAMES_PROPERTY.keys())

        if sum(k is not None for k in key) == 0:
            break
        # 'color': sheet.cell(row=y, column=2).fill.start_color.rgb

        group = str(sheet[f'{COL_GROUP}{y}'].value).lower()
        if 'oборудование' in group:
            continue

        count_value = str(sheet[f'{COL_COUNT}{y}'].value).split()
        if len(count_value) > 1:
            count, unit = count_value
        else:
            count, unit = count_value[0], 'шт.'

        count = float(count.replace(',', '.'))
        value = {
            'count': count,
            'unit': unit,
            'after': 0
            }

        if key in dict_current_sheet:
            dict_current_sheet[key]['count'] += count
        else:
            dict_current_sheet[key] = value
        
        if dict_sheet_after:
            if key in dict_sheet_after:
                dict_current_sheet[key]['after'] = dict_sheet_after[key]['count']
            else:
                dict_current_sheet[key]['after'] = None

    return dict_current_sheet

def create_preprocess_sheet(sheet_name: str, book: Workbook, book_path: str, dict_sheet_after: dict=None) -> dict:
    """Создание на основе get_dict_from_sheet нового листа"""
    sheet = book[sheet_name]
    dict_sheet = get_dict_from_sheet(sheet=sheet, dict_sheet_after=dict_sheet_after)
    new_sheet = book.create_sheet(f'_{sheet.title}')

    for x, title_column in enumerate(COL_NAMES_PROPERTY.values(), 1):
        new_sheet.cell(row=1, column=x).value = title_column
    new_sheet.cell(row=1, column=len(COL_NAMES_PROPERTY) + 1).value = 'Количество'
    new_sheet.cell(row=1, column=len(COL_NAMES_PROPERTY) + 2).value = 'Единичная величина'

    for y, (key, value) in enumerate(dict_sheet.items(), 2):
        for x, prop in enumerate(key, 1):
            new_sheet.cell(row=y, column=x).value = prop
        new_sheet.cell(row=y, column=len(key) + 1).value = value['count']
        new_sheet.cell(row=y, column=len(key) + 2).value = value['unit']
        
        if value['after'] is None:
            new_sheet.cell(row=y, column=len(key) + 3).value = 'new'
        else:
            diff = value['count'] - value['after']
            if diff != 0:
                new_sheet.cell(row=y, column=len(key) + 3).value = diff

    book.save(book_path)
    return dict_sheet



def proccesing_last_sheet(book: Workbook, book_path: str) -> tuple:
    sheet_names = book.sheetnames
    current_sheet = book[sheet_names[-2]]
    last_sheet = book[sheet_names[-1]]

    max_count_column = 12
    using_column_name = [value for x in range(1, max_count_column) if (value := current_sheet.cell(row=1, column=x).value) is not None]
    
    for x in range(1, max_count_column):
        value = last_sheet.cell(row=1, column=x).value
        if value not in using_column_name:
            last_sheet.delete_cols(x)
    
    dict_current_sheet = get_dict_from_sheet(sheet=current_sheet)

    for y in range(2, 1000):
        key = tuple(last_sheet[f'{col_adrs}{y}'].value for col_adrs in COL_NAMES_PROPERTY) + (last_sheet[f'{COL_MATERIAL}{y}'].value, )

        if sum(k is not None for k in key) == 0:
            break
        
        if key in dict_current_sheet:
            color = dict_current_sheet[key]['color']
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for cell in last_sheet[y]:
                cell.fill = fill


    book.save(book_path)



 
def main() -> None:
    path = r'\\PDM\PKODocs\Inventor Project\ООО ЛебедяньМолоко\1648_25\4.1. УРП ALS.P.01.06.02.A\04 Исходящие\ALS.1648.4.1.01. Из инвентора.xlsx'

    name_sheet_after = '16.10.2025'
    name_sheet_before = '31.10.2025'

    book: Workbook = load_workbook(path)
    
    # proccesing_last_sheet(book, path)
    dict_sheet_after = create_preprocess_sheet(sheet_name=name_sheet_after, book=book, book_path=path)
    dict_sheet_before = create_preprocess_sheet(sheet_name=name_sheet_before, book=book, book_path=path, dict_sheet_after=dict_sheet_after)


if __name__ == '__main__':
    main()
