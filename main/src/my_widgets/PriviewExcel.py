import sys
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets, Qt


class Tab(QtWidgets.QWidget):
    signal_change_sheet = QtCore.pyqtSignal(str)

    def __init__(self, parent, sheet, sheet_name: str):
        super().__init__(parent)
        self.sheet = sheet
        self.sheet_name = sheet_name
        self.init()

    def init(self):
        self.vertivalLayout = QtWidgets.QVBoxLayout(self)
        self.setObjectName(self.sheet_name)

        self.table = QtWidgets.QTableWidget(self)
        self.table.setStyleSheet("""
        QTableWidget {
            border: 1px solid black
        }
        QTableView::item {
            border: 1px solid black
        }
        """)
        self.vertivalLayout.addWidget(self.table)
        self.fill_table()

    def fill_table(self):
        self.table.setRowCount(self.sheet.max_row)
        self.table.setColumnCount(self.sheet.max_column)

        for rng in self.sheet.merged_cells.ranges:
            start_cell, *_, end_cell = list(rng.cells)
            start_row, start_col = start_cell
            end_row, end_col = end_cell
            self.table.setSpan(start_row - 1, start_col - 1, end_row - start_row + 1, end_col - start_col + 1)

        for y in range(self.sheet.max_row):
            for x in range(self.sheet.max_column):
                cell = self.sheet.cell(y + 1, x + 1)
                item = str(cell.value)
                if str(item) != 'None':
                    qItem = QtWidgets.QTableWidgetItem(str(item))
                    self.table.setItem(y, x, qItem)
                    qItem.setFlags(QtCore.Qt.ItemIsEnabled)


class PreviewExcelFile(QtWidgets.QWidget):
    signal_load_file = QtCore.pyqtSignal(str)
    signal_change_tab = QtCore.pyqtSignal(str)

    def __init__(self, parent, filepath):
        super().__init__(parent)
        self.filepath = filepath
        self.init()

    def init(self):
        self.centralGridLayout = QtWidgets.QGridLayout(self )

        self.tab_widget = QtWidgets.QTabWidget(self)
        self.tab_widget.currentChanged.connect(self.eventChangeTab)
        self.centralGridLayout.addWidget(self.tab_widget)

        # self.signal_load_file.connect(self.read_file)
        self.read_file(self.filepath)

    def load_file(self):
        dlg = QtWidgets.QFileDialog()
        # dlg.setFileMode(QtWidgets.QFileDialog.ExistingFiles)
        file_format_filter = '*.xls, *.xlsx'
        dlg.setNameFilter(f"Excel ({file_format_filter})")
        dlg.selectNameFilter(f"Excel ({file_format_filter})")
        dlg.setWindowTitle('Выберете файл')
        dlg.exec_()

        filepath = dlg.selectedFiles()
        if filepath:
            self.signal_load_file.emit(filepath[0])

    def read_file(self, pathfile: str):
        book: openpyxl.Workbook = openpyxl.load_workbook(pathfile, data_only=True)
        for sheen_name in book.sheetnames:
            sheet = book[sheen_name]
            tab = Tab(parent=self, sheet=sheet, sheet_name=sheen_name)
            self.tab_widget.addTab(tab, sheen_name)

    def currentTab(self) -> str:
        return self.tab_widget.currentWidget().objectName()

    def eventChangeTab(self, event) -> None:
        self.signal_change_tab.emit(self.currentTab())


def my_excepthook(type, value, tback):
    QtWidgets.QMessageBox.critical(
        window, "CRITICAL ERROR", str(value),
        QtWidgets.QMessageBox.Cancel
    )


if __name__ == '__main__':
    sys.excepthook = my_excepthook

    app = QtWidgets.QApplication(sys.argv)

    window = PreviewExcelFile()
    window.show()
    sys.exit(app.exec_())
