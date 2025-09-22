"""
Получение данных об изделиях из
*Autocad - схема;
*Excel - спецификация
"""

import win32com.client as wc
import pythoncom
import openpyxl

try:
    from ..function.GetFromDB import query_get_db, query_lst_to_str
    from ..function.HashingArt import my_hash
except ImportError:
    from GetFromDB import query_get_db, query_lst_to_str
    from main.src.function.HashingArt import my_hash


def decorated_try(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception:
            pass
    return wrapper


def get_selection_screen(signal=None) -> dict:
    """
    Получение данных из схемы в Autocad. Парсинг свойств блоков
    :param signal: сигнал для processbar PyQt
    :return: {"Таг номер": ["Артикул1", "Артикул2", ...], ...}
    """
    pythoncom.CoInitialize()

    app = wc.GetActiveObject('Autocad.Application.24')
    document = app.ActiveDocument

    name_selection = 'tmp'
    try:
        document.SelectionSets(name_selection).Delete()
    except Exception:
        pass

    selection = document.SelectionSets.Add(name_selection)
    selection.SelectOnScreen()

    attrs = {}
    title = None
    count = 0
    while True:
        try:
            sel_sets = document.SelectionSets(name_selection)
            one = 100 / len(sel_sets)
            counter = 0
            for obj in sel_sets:
                if signal is not None:
                    signal.emit(int(counter))
                    counter += one

                name = obj.ObjectName
                if name == 'AcDbBlockReference':
                    attr = []
                    for i in obj.GetAttributes():
                        tag = i.TagString
                        text = i.TextString
                        if 'PROD' in tag and len(text) > 0:
                            attr.append(text)
                        if tag == 'TAG_N':
                            text = text if text else 'Нет TAG номер'
                            title = (text, obj.EffectiveName)
                            attrs[title] = None

                    if attr:
                        attrs[title] = attr
            break
        except Exception as error:
            count += 1
            if count > 10:
                return {}
            continue

    return attrs


def get_from_excel(filepath: str, sheet_name=None, signal=None) -> dict:
    """
    Функция для того, чтобы получать словарь по структуре схожей со структурой из ф-ции get_selection_screen
    Но так как на листе с нумерацией нет такгов, то таги берутся с листа с тагами и по артикулам находятся
    на необходимом листе (sheet_name), чтобы составить словарь с нумерацией

    :param filepath: путь к таговой спецификации
    :param sheet_name: имя листа с нумерацией
    :param signal: сигнал для processbar PyQt
    :return: {"Таг номер": {"Артикул:1 ["Информация"], "Артикул2": [...]"}, ...}
    """
    doc: openpyxl.Workbook = openpyxl.load_workbook(filepath)

    blocks_number = {}
    if sheet_name is not None:
        sheet = doc[sheet_name]
        numbers = sheet[f'A8:A{sheet.max_row}']
        tags = sheet[f'B8:B{sheet.max_row}']

        for n, t in zip(numbers, tags):
            blocks_number[t[0].value] = n[0].value

    sheet = doc['Спецификация с TAG-номерами']

    tags = sheet[f'G8:G{sheet.max_row}']
    arts = sheet[f'J8:J{sheet.max_row}']
    naming = sheet[f'K8:K{sheet.max_row}']
    des = sheet[f'L8:L{sheet.max_row}']
    manufacturer = sheet[f'P8:P{sheet.max_row}']

    blocks = {}
    for t, *info in zip(tags, arts, naming, des, manufacturer):
        if any(i[0].value is not None for i in info):
            t = t[0].value
            info = [i[0].value for i in info]

            art = info[0]
            number = blocks_number.get(art)
            if t not in blocks:
                blocks[t] = {art: [number] + info[1:]}
            else:
                blocks[t].update({info[0]: [number] + info[1:]})

    return blocks


def block_dict(blocks: dict, signal=None) -> dict:
    """
    Поиск в БД изделий по артикулам и выгрузка необходимых полей
    :param blocks: структура из функции get_selection_screen
    :param signal: сигнал для processbar PyQt
    :return: {("ТАГ номер", "Имя блока"): {"Артикул1" [Инфа], ... , "filepath"}: "путь к детали или сборки в Inventor"}
    """
    data = {}
    for tag_n, block in blocks.items():
        if block is None:
            continue

        lst = tuple(block) + ('',)

        query = f"SELECT artikul, description, characteristics, manufacturer FROM vw__fullprod WHERE artikul in {query_lst_to_str(lst)}"
        res_db = query_get_db(query)

        value = {res[0]: (None, *res[1:]) for res in res_db}

        details_hash = my_hash(block)
        query = f"""SELECT filepath FROM dbo.co_detail WHERE hash='{details_hash}'"""
        filepath = query_get_db(query)

        value['filepath'] = filepath
        data[tag_n] = value

    if signal is not None:
        signal.emit(100)

    return data


if __name__ == '__main__':
    p = r'C:\Users\p.golubev\Desktop\Проекты\Мечта ООО\9.1 ALS.P.01.05.05.A\Инфа\Зона 206 Отделение приготовления смесей 24.08.23.xlsx'
    dct = get_selection_screen()
    # print(dct)
    # dct = get_from_excel(p, sheet_name='6.9')

    d = block_dict(dct)
    print(d)
    # for item in dct.items():
    #     print(item)


